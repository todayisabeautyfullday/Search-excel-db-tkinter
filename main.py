import openpyxl
import tkinter as tk
from tkinter import filedialog

def merge_excel_sheets(files):
    # Create a new Workbook
    merged_wb = openpyxl.Workbook()
    merged_sheet = merged_wb.active
    
    # Counter for the row in the merged sheet
    row = 1
    
    # Loop through each file
    for file in files:
        # Load the Workbook
        wb = openpyxl.load_workbook(file)
        # Get the first sheet
        sheet = wb.active
        
        # Get the number of columns in the current sheet
        num_cols = sheet.max_column
        
        # Loop through each row in the sheet
        for r in sheet.iter_rows(values_only=True):
            # Write the values from the current row to the merged sheet
            for col, value in enumerate(r[:num_cols], 1):
                merged_sheet.cell(row=row, column=col, value=value)
            row += 1
    
    return merged_wb

def search_database(wb, search_term):
    sheet = wb.active
    
    # Get the number of columns in the sheet
    num_cols = sheet.max_column
    
    # Check each cell in the first column
    for row in sheet.iter_rows(values_only=True):
        if row[0] == search_term:
            return row[:num_cols]
    return None

def choose_files():
    files = filedialog.askopenfilenames(title="Select Excel Files", filetypes=[("Excel Files", "*.xlsx")])
    return files

def search():
    search_term = search_entry.get()
    result = search_database(merged_wb, search_term)
    result_label.config(text=result)

# Create the GUI
root = tk.Tk()
root.title("Excel Database")

# Create the Choose Files button
choose_files_button = tk.Button(root, text="Choose Files", command=choose_files)
choose_files_button.pack()

# Create the Search Entry
search_entry = tk.Entry(root)
search_entry.pack()

# Create the Search button
search_button = tk.Button(root, text="Search", command=search)
search_button.pack()

# Create the Result Label
result_label = tk.Label(root, text="")
result_label.pack()

# Merge the Excel sheets
files = choose_files()
merged_wb = merge_excel_sheets(files)

# Start the GUI
root.mainloop()
